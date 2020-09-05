#-*- encoding: utf-8 -*-
import xml.etree.ElementTree as etree
import os
import xlsxwriter
import sys
from PySide2.QtCore import Qt
from PySide2 import QtGui, QtCore, QtWidgets
from PySide2.QtWidgets import QApplication, QMainWindow, QFileDialog, QWidget, QTableWidget,QTableWidgetItem,QPushButton,QListView,QAbstractItemView,QTreeView,QMessageBox
from PySide2.QtCore import QFile, QRect
from PySide2.QtGui import QIcon
from gui import Ui_MainWindow
import json
from os.path import dirname, realpath, join, abspath
from openpyxl.styles import Alignment

from openpyxl import load_workbook,  Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.datavalidation import DataValidation
from FacturasLocal import FacturaLocal as Factura

# pyside2-uic gui.ui -o gui.py
#C:\Users\arabela\Anaconda3\Scripts\pyinstaller --noconsole viaticos.spec
#excludes=['scipy','numpy']
#C:\Users\arabela\Anaconda3\Scripts\pyinstaller viaticos.spec
class EditPersonasDialog(QWidget):
    def __init__(self):
        QWidget.__init__(self)

        try:
            self.dirPath = dirname(abspath(__file__))
        except NameError:  # We are the main py2exe script, not a module
            self.dirPath = dirname(abspath(sys.argv[0]))



class MainWindow(QMainWindow):
    def __init__(self):
        super(MainWindow, self).__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)

        self.ui.input_carpeta.clicked.connect(self.cualCarpeta)
        self.ui.carpeta_personal.clicked.connect(self.prellenar_formato)
        self.ui.excel_button.clicked.connect(self.abre_excel)
        try:
            self.dirPath = dirname(abspath(__file__))
        except NameError:  # We are the main py2exe script, not a module
            self.dirPath = dirname(abspath(sys.argv[0]))

        self.ui.tableWidget.setColumnCount(6)
        self.ui.tableWidget.setColumnWidth(0,120)#viaje
        self.ui.tableWidget.setColumnWidth(1,70)#total
        self.ui.tableWidget.setColumnWidth(2,70)#transporte
        self.ui.tableWidget.setColumnWidth(3,70)#hospedaje
        self.ui.tableWidget.setColumnWidth(4,70)#alimentos
        self.ui.tableWidget.setColumnWidth(5,70)#otros
        #header = self.tableWidget.verticalHeader()
        self.ponEncabezado()
        self.dicc_users = {}
        self.dicc_viajes = {}

    def esteItem(self, text, tooltip):
        item = QTableWidgetItem(text)
        item.setToolTip(tooltip)
        item.setFlags(item.flags() ^ Qt.ItemIsEditable)
        return item
    def tipo_de_gasto(self, clave_ps):
        tipo = "Otros"
        if clave_ps.startswith('7811') or clave_ps.startswith('9511') or clave_ps.startswith('1510') or clave_ps.startswith('1511') or clave_ps.startswith('1512'):
            tipo = "Transporte"
        if clave_ps.startswith('9010'):
            tipo = "Alimentos"
        if clave_ps.startswith('9011'):
            tipo = "Hospedaje"

        return(tipo)


    def abre_excel(self):
        try:
            os.startfile(self.xlsx_path)
        except:
            QMessageBox.information(self, "Information", "El sistema no tiene una aplicación por default para abrir exceles" )

    def prellenar_formato(self):
        self.ui.textBrowser.clear()
        self.ui.tableWidget.clear()
        self.ui.tableWidget.repaint()
        self.ponEncabezado()
        self.ui.tableWidget.setRowCount(6)
        self.ui.tableWidget.repaint()
        file_dialog = QFileDialog()
        file_dialog.setFileMode(QFileDialog.DirectoryOnly)
        file_dialog.setOption(QFileDialog.DontUseNativeDialog, True)
        file_view = file_dialog.findChild(QListView, 'listView')
        if file_dialog.exec():
            path = file_dialog.selectedFiles()[0]

        empleado_folder = os.path.split(path)[1]

        viaje_folder = os.path.split(os.path.split(path)[0])[1]


        home = os.path.expanduser('~')
        template_folder = os.path.join(home, 'Documents', 'huiini')
        wb_template = load_workbook(os.path.join(template_folder,"FORMATO_template.xlsx"))
        ws_template = wb_template[wb_template.get_sheet_names()[0]]


        self.xlsx_path = os.path.join(path,"FORMATO_"+empleado_folder.replace(" ","-")+"_"+viaje_folder.replace(" ","-")+".xlsx")


        row_f = 11
        #folder = os.path.split(path)[1]
        por_tipo={"Total":0,"Transporte":0,"Alimentos":0,"Hospedaje":0,"Otros":0}
        hay_facturas = False
        for archivo in os.listdir(path):
            if archivo.endswith(".xml"):
                row_f += 1
                xml_path = os.path.join(path,archivo)
                factura = Factura(xml_path)
                total = factura.total
                iva = factura.traslados["IVA"]["importe"]
                uuid = factura.UUID
                fecha = factura.fechaTimbrado.split("T")[0]
                provedor = factura.EmisorNombre
                clave_ps = factura.conceptos[0]['clave_concepto']
                tipo = self.tipo_de_gasto(clave_ps)
                ws_template.cell(row_f, 1,    fecha)#fecha
                ws_template.cell(row_f, 2,    uuid)#fecha
                ws_template.cell(row_f, 3,    tipo)#tipo
                ws_template.cell(row_f, 4,    provedor)#provedor
                ws_template.cell(row_f, 5,    factura.subTotal )#importe
                ws_template.cell(row_f, 6,    iva)#traslado_iva
                ws_template.cell(row_f, 7,    total)
                por_tipo[tipo]+=total
                por_tipo["Total"]+=total
                hay_facturas = True


        if hay_facturas:
            self.ui.textBrowser.append("Procesando: "+empleado_folder+" "+viaje_folder)
            self.ui.tableWidget.setItem(1,0,self.esteItem(empleado_folder,""))
            self.ui.tableWidget.setItem(1,1,self.esteItem(str(por_tipo["Total"]),""))
            self.ui.tableWidget.setItem(1,2,self.esteItem(str(por_tipo["Transporte"]),""))
            self.ui.tableWidget.setItem(1,3,self.esteItem(str(por_tipo["Hospedaje"]),""))
            self.ui.tableWidget.setItem(1,4,self.esteItem(str(por_tipo["Alimentos"]),""))
            self.ui.tableWidget.setItem(1,5,self.esteItem(str(por_tipo["Otros"]),""))
            try:
                mes = viaje_folder.split(" ")[1].title()
                dia = viaje_folder.split(" ")[0]
            except:
                mes = ""
                dia = ""
            ws_template.cell(5, 7, mes)#dia
            ws_template.cell(5, 5, dia)#mes
            self.ui.textBrowser.append("Formato de gastos de viaje creado en: "+self.xlsx_path)
            wb_template.save(self.xlsx_path)
            self.ui.excel_button.setEnabled(True)
        else:
            QMessageBox.information(self, "Information", "No hay facturas en este folder" )

    def ponEncabezado(self):
        self.ui.tableWidget.setHorizontalHeaderItem (0, QTableWidgetItem("Viaje"))
        self.ui.tableWidget.setHorizontalHeaderItem (1, QTableWidgetItem("Total"))
        self.ui.tableWidget.setHorizontalHeaderItem (2, QTableWidgetItem("Transporte"))
        self.ui.tableWidget.setHorizontalHeaderItem (3, QTableWidgetItem("hospedaje"))
        self.ui.tableWidget.setHorizontalHeaderItem (4, QTableWidgetItem("Alimentos"))
        self.ui.tableWidget.setHorizontalHeaderItem (5, QTableWidgetItem("Otros"))

    def cualCarpeta(self):
        file_dialog = QFileDialog()
        file_dialog.setFileMode(QFileDialog.DirectoryOnly)
        file_dialog.setOption(QFileDialog.DontUseNativeDialog, True)
        file_view = file_dialog.findChild(QListView, 'listView')

        # to make it possible to select multiple directories:
        if file_view:
            file_view.setSelectionMode(QAbstractItemView.MultiSelection)
        f_tree_view = file_dialog.findChild(QTreeView)
        if f_tree_view:
            f_tree_view.setSelectionMode(QAbstractItemView.MultiSelection)

        if file_dialog.exec():
            paths = file_dialog.selectedFiles()

        self.procesaCarpetas(paths)


    def procesa_formato(self,excel_path,usuario,vieje_folder):
        fromato_wb = load_workbook(excel_path, data_only=True)
        ws = fromato_wb[fromato_wb.get_sheet_names()[0]]

        for f_row in range(12,47):
            total = ws.cell(f_row,7).value
            if not total == None:
                self.row += 1
                fecha = ws.cell(f_row,1).value
                nombre = ws.cell(f_row,2).value
                tipo = ws.cell(f_row,3).value
                self.dicc_users[usuario] += float(total)
                self.dicc_viajes[vieje_folder]["Total"] += total
                self.dicc_viajes[vieje_folder][tipo] += total
                self.worksheet.cell(self.row, 1,     usuario)
                self.worksheet.cell(self.row, 2,     vieje_folder)
                cell_fecha = self.worksheet.cell(self.row, 3)
                cell_fecha.value = fecha
                cell_fecha.number_format = 'dd/mm/YYYY'
                self.worksheet.cell(self.row, 4,     nombre)
                #################worksheet.data_validation('E'+str(row+1), {'validate': 'list', 'source': ['Alimentos', 'Hospedaje', 'Transporte', 'Otros']})
                self.worksheet.cell(self.row, 5,     tipo)
                self.worksheet.cell(self.row, 6,     total)


    def procesaCarpetas(self,paths):
        self.ui.textBrowser.clear()
        self.ui.tableWidget.clear()
        self.ui.tableWidget.repaint()
        self.ponEncabezado()
        self.ui.tableWidget.setRowCount(6)
        self.ui.tableWidget.repaint()
        head = os.path.split(paths[0])[0]
        tail = os.path.split(head)[1]

        #print(head)

        self.xlsx_path = os.path.join(head,"resumen_"+tail+".xlsx")
        if os.path.isfile(self.xlsx_path):
            workbook = load_workbook(self.xlsx_path)
            ws_empleados = workbook["Empleados"]
            ws_viajes = workbook["Viajes"]
        else:
            workbook = Workbook()
            ws_empleados = workbook.create_sheet("Empleados")
            ws_viajes = workbook.create_sheet("Viajes")
            ws_viajes.cell(1, 2, "Anticipo")
            ws_viajes.cell(1, 3, "Total")
            ws_viajes.cell(1, 4, "Transporte")
            ws_viajes.cell(1, 5, "Hospedaje")
            ws_viajes.cell(1, 6, "Alimentos")
            ws_viajes.cell(1, 7, "Otros")
            ws_viajes.column_dimensions[get_column_letter(1)].width = 17
            ws_viajes.column_dimensions[get_column_letter(2)].width = 10
            ws_viajes.column_dimensions[get_column_letter(3)].width = 10
            ws_viajes.column_dimensions[get_column_letter(4)].width = 10
            ws_viajes.column_dimensions[get_column_letter(5)].width = 10
            ws_viajes.column_dimensions[get_column_letter(6)].width = 10



            sheet1_name = workbook.get_sheet_names()[0]
            sheet1 = workbook[sheet1_name]
            workbook.remove_sheet(sheet1)
            ws_empleados.column_dimensions[get_column_letter(1)].width = 24
            ws_empleados['B1'].alignment=Alignment(horizontal='center',vertical='center',text_rotation=0,wrap_text=True,shrink_to_fit=False,indent=0)
            ws_empleados.column_dimensions[get_column_letter(2)].width = 12
            ws_empleados.cell(1, 2, "ANTICIPO")
            ws_empleados['C1'].alignment=Alignment(horizontal='center',vertical='center',text_rotation=0,wrap_text=True,shrink_to_fit=False,indent=0)
            # ws_empleados['C1'].alignment = Alignment(wrap_text=True)
            # ws_empleados["C1"].alignment.vertical = "center"
            ws_empleados['C1'].value = "TOTAL\nGASTO"
            ws_empleados.column_dimensions[get_column_letter(4)].width = 17
            ws_empleados['D1'].alignment=Alignment(horizontal='center',vertical='center',text_rotation=0,wrap_text=True,shrink_to_fit=False,indent=0)
            ws_empleados['D1'].value = "SALDO\nPARA\nMETROPOLITANA"
            ws_empleados.column_dimensions[get_column_letter(5)].width = 16
            ws_empleados['E1'].alignment=Alignment(horizontal='center',vertical='center',text_rotation=0,wrap_text=True,shrink_to_fit=False,indent=0)
            ws_empleados['E1'] = "SALDO\nA FAVOR\nEMPLEADO"
            ws_empleados['F1'].alignment=Alignment(horizontal='center',vertical='center',text_rotation=0,wrap_text=True,shrink_to_fit=False,indent=0)
            ws_empleados.cell(1, 6, "BALANCE")
            #FFE082
            ws_empleados.cell(1, 2).fill = PatternFill(start_color="FFE082", fill_type = "solid")
            ws_empleados.cell(1, 3).fill = PatternFill(start_color="FFE082", fill_type = "solid")
            ws_empleados.cell(1, 4).fill = PatternFill(start_color="FFE082", fill_type = "solid")
            ws_empleados.cell(1, 5).fill = PatternFill(start_color="FFE082", fill_type = "solid")
            ws_empleados.cell(1, 6).fill = PatternFill(start_color="FFE082", fill_type = "solid")
            c = ws_empleados['A2']
            ws_empleados.freeze_panes = c

        self.ui.textBrowser.append("Carpeta: "+ head)
        lista_carpetas_personales = []
        self.dicc_viajes = {}
        for viaje in paths:
            vieje_folder = os.path.split(viaje)[1]
            if vieje_folder not in workbook.sheetnames:

                self.worksheet = workbook.create_sheet(vieje_folder)

                self.worksheet.cell(1, 1,     "Usuario")
                self.worksheet.cell(1, 2,     "Viaje")
                self.worksheet.cell(1, 3,     "Fecha")
                self.worksheet.cell(1, 4,     "Folio facturas")
                self.worksheet.cell(1, 5,     "Tipo")
                self.worksheet.cell(1, 6,     "Monto")
                self.worksheet.column_dimensions[get_column_letter(1)].width = 23
                self.worksheet.column_dimensions[get_column_letter(2)].width = 15
                self.worksheet.column_dimensions[get_column_letter(3)].width = 12
                self.worksheet.column_dimensions[get_column_letter(5)].width = 10
                self.worksheet.column_dimensions[get_column_letter(7)].width = 10
                self.worksheet.column_dimensions[get_column_letter(12)].width = 10
                self.worksheet.column_dimensions[get_column_letter(13)].width = 10
                self.worksheet.column_dimensions[get_column_letter(14)].width = 10
                self.worksheet.column_dimensions[get_column_letter(9)].width = 23

                self.row = 1
                suma = 0
                arch_total=0
                self.dicc_users = {}
                self.dicc_viajes[vieje_folder]={"Total":0,"Transporte":0,"Alimentos":0,"Hospedaje":0,"Otros":0}
                for carpeta in os.listdir(viaje):
                    if os.path.isdir(os.path.join(viaje,carpeta)):
                        carpeta_usuario = os.path.join(viaje,carpeta)

                        #carpeta_usuario = os.path.join(path,carpeta)
                        usuario = os.path.split(carpeta_usuario)[1]
                        head = os.path.split(os.path.split(carpeta_usuario)[0])[1]
                        #print(head)
                        #viaje = carpeta_usuario.split("")
                        self.ui.textBrowser.append("Procesando: "+ usuario)
                        suma_user=0
                        self.dicc_users[usuario]=0
                        excel_path = ""
                        for archivo in os.listdir(carpeta_usuario):
                            if archivo.endswith(".xlsx") and archivo.startswith("FORMATO"):
                                excel_path = os.path.join(carpeta_usuario,archivo)
                                #self.procesa_formato(self,excel_path,usuario,vieje_folder)
                                self.procesa_formato(excel_path,usuario,vieje_folder)
                        if excel_path == "":
                            for archivo in os.listdir(carpeta_usuario):
                                if archivo.endswith(".xml"):
                                    arch_total+=1
                                    self.row += 1
                                    xml_path = os.path.join(carpeta_usuario,archivo)
                                    factura = Factura(xml_path)
                                    total = factura.total
                                    iva = factura.traslados["IVA"]["importe"]
                                    uuid = factura.UUID
                                    fecha = factura.fechaTimbrado.split("T")[0]
                                    provedor = factura.EmisorNombre
                                    clave_ps = factura.conceptos[0]['clave_concepto']
                                    tipo = self.tipo_de_gasto(clave_ps)
                                    self.dicc_users[usuario] += total
                                    self.dicc_viajes[vieje_folder]["Total"] += total
                                    self.dicc_viajes[vieje_folder][self.tipo_de_gasto(clave_ps)] += total

                                    self.worksheet.cell(self.row, 1,     usuario)
                                    self.worksheet.cell(self.row, 2,     head)
                                    self.worksheet.cell(self.row, 3,     fecha)
                                    self.worksheet.cell(self.row, 4,     uuid)
                                    #################self.worksheet.data_validation('E'+str(row+1), {'validate': 'list', 'source': ['Alimentos', 'Hospedaje', 'Transporte', 'Otros']})
                                    self.worksheet.cell(self.row, 5,     tipo)
                                    self.worksheet.cell(self.row, 6,     total)
                                    suma+=total

                sumRow = self.row + 1
                self.worksheet.cell(sumRow, 5,     "Suma")
                self.worksheet.cell(sumRow, 6,     '=SUM(F2:F'+str(self.row)+')')
                #sumas_row = sumRow + 2
                #sumas_row_inicial = str(sumRow + 2)
                sumas_row = 2
                sumas_row_inicial = "2"
                sumas_column_inicial = 10
                suma_total = 0
                #format = workbook.add_format()

                #format.set_pattern(1)
                #format.set_bg_color('#d3d3d3')

                self.worksheet.column_dimensions[get_column_letter(1)].width = 24
                self.worksheet['J1'].alignment=Alignment(horizontal='center',vertical='center',text_rotation=0,wrap_text=True,shrink_to_fit=False,indent=0)
                self.worksheet.column_dimensions[get_column_letter(10)].width = 12
                self.worksheet.cell(1, 10, "ANTICIPO")
                self.worksheet['K1'].alignment=Alignment(horizontal='center',vertical='center',text_rotation=0,wrap_text=True,shrink_to_fit=False,indent=0)
                self.worksheet['K1'].value = "TOTAL\nGASTO"
                self.worksheet.column_dimensions[get_column_letter(11)].width = 12
                self.worksheet['L1'].alignment=Alignment(horizontal='center',vertical='center',text_rotation=0,wrap_text=True,shrink_to_fit=False,indent=0)
                self.worksheet['L1'].value = "SALDO\nPARA\nMETROPOLITANA"
                self.worksheet.column_dimensions[get_column_letter(12)].width = 16
                self.worksheet['M1'].alignment=Alignment(horizontal='center',vertical='center',text_rotation=0,wrap_text=True,shrink_to_fit=False,indent=0)
                self.worksheet['M1'] = "SALDO\nA FAVOR\nEMPLEADO"
                self.worksheet.column_dimensions[get_column_letter(13)].width = 15
                self.worksheet['N1'].alignment=Alignment(horizontal='center',vertical='center',text_rotation=0,wrap_text=True,shrink_to_fit=False,indent=0)
                self.worksheet['N1'] = "Estatus"
                self.worksheet.cell(1, 10).fill = PatternFill(start_color="FFE082", fill_type = "solid")
                self.worksheet.cell(1, 11).fill = PatternFill(start_color="FFE082", fill_type = "solid")
                self.worksheet.cell(1, 12).fill = PatternFill(start_color="FFE082", fill_type = "solid")
                self.worksheet.cell(1, 13).fill = PatternFill(start_color="FFE082", fill_type = "solid")
                self.worksheet.cell(1, 14).fill = PatternFill(start_color="FFE082", fill_type = "solid")
                dv = DataValidation(type="list", formula1='"Pendiente,Pagado"', allow_blank=True)
                self.worksheet.add_data_validation(dv)

                for key, value in self.dicc_users.items():
                    self.ui.textBrowser.append(key+": "+ str(value))
                    self.worksheet.cell(sumas_row, sumas_column_inicial-1, key)
                    #worksheet.write(sumas_row, 5,     value)
                    self.worksheet.cell(sumas_row, sumas_column_inicial+1,     '=SUMIF(A2:A'+str(self.row)+',"'+key+'",F2:F'+str(self.row)+')')

                    self.worksheet.cell(sumas_row, sumas_column_inicial+2,     '=IF(N'+str(sumas_row)+'="Pagado",0,IF(J'+str(sumas_row)+'-K'+str(sumas_row)+'>0,J'+str(sumas_row)+'-K'+str(sumas_row)+',0))')#
                    self.worksheet.cell(sumas_row, sumas_column_inicial+3,     '=IF(N'+str(sumas_row)+'="Pagado",0,IF(K'+str(sumas_row)+'-J'+str(sumas_row)+'>0,K'+str(sumas_row)+'-J'+str(sumas_row)+',0))')

                    dv.add(self.worksheet.cell(sumas_row, sumas_column_inicial+4))
                    self.worksheet.cell(sumas_row, sumas_column_inicial+4,     'Pendiente')

                    suma_total += value
                    sumas_row += 1
                    ya_estaba = False
                    for row1 in ws_empleados.iter_rows(min_row=1, max_col=1, max_row=len(ws_empleados['A'])):
                        for cell in row1:
                            if cell.value == key:
                                user_first_row = cell.row
                                ya_estaba = True
                                print("ya estaba "+key)

                    data_row = 0
                    if ya_estaba:
                        #if name on resumen then last row of the Name and insert a row

                        for row1 in ws_empleados.iter_rows(min_row=1, max_col=1, max_row=len(ws_empleados['A'])+1):
                            #print(row1[0].value,row1[0].row)
                            if (row1[0].row > user_first_row) and (row1[0].value == None):
                                data_row = row1[0].row
                                break
                        ws_empleados.insert_rows(data_row)

                    else:
                        #else last row of document + 2 and add Name
                        ultima = 0
                        for kk in ws_empleados['A']:
                            if not kk == None:
                                ultima = kk.row
                        if ultima == 2:
                            user_first_row = 2
                        else:
                            user_first_row = ultima + 3
                        print(key,user_first_row,ultima)
                        ws_empleados.cell(user_first_row, 1, key)

                        ws_empleados.cell(user_first_row, 1).fill = PatternFill(start_color="C8E6C9", fill_type = "solid")
                        ws_empleados.cell(user_first_row, 1).font = Font(bold=True)
                        data_row = user_first_row+1
                        print("no estaba "+key)

                    print("data_row = "+str(data_row))
                    ws_empleados.cell(data_row, 1, vieje_folder)
                    anticipo_str = "='"+vieje_folder+"'!J"+str(sumas_row-1)
                    #print("anticipo = " + anticipo_str)
                    ws_empleados.cell(data_row, 2, anticipo_str)
                    ws_empleados.cell(data_row, 3, value)

                    resta_metro_str = "'"+vieje_folder+"'!J"+str(sumas_row-1)+"-'"+vieje_folder+"'!K"+str(sumas_row-1)
                    saldo_metro = "=IF("+resta_metro_str+">0,"+resta_metro_str+",0)"
                    ws_empleados.cell(data_row, 4, saldo_metro)#saldo metro
                    resta_empleado_str = "'"+vieje_folder+"'!K"+str(sumas_row-1)+"-'"+vieje_folder+"'!J"+str(sumas_row-1)
                    saldo_empleado = "=IF("+resta_empleado_str+">0,"+resta_empleado_str+",0)"
                    ws_empleados.cell(data_row, 5, saldo_empleado)#saldo empleado


                totales_row_str = str(sumas_row-1)

                self.worksheet.cell(sumas_row, sumas_column_inicial,     '=SUM(J'+sumas_row_inicial+':J'+totales_row_str+')')
                self.worksheet.cell(sumas_row, sumas_column_inicial+1,     '=SUM(K'+sumas_row_inicial+':K'+totales_row_str+')')
                self.worksheet.cell(sumas_row, sumas_column_inicial+2,     '=SUM(L'+sumas_row_inicial+':L'+totales_row_str+')')
                self.worksheet.cell(sumas_row, sumas_column_inicial+3,     '=SUM(M'+sumas_row_inicial+':M'+totales_row_str+')')


                self.ui.textBrowser.append("Total: "+str(suma_total))
            else:
                self.ui.textBrowser.append("La carpeta de viaje "+vieje_folder+" ya estaba procesada")
        #recalcular las sumas por persona
        anterior_is_number = False
        for cell in ws_empleados['B']:
            termino = False
            if cell.row > 1:

                if not anterior_is_number and cell.value != None:
                    empieza_empleado = cell.row
                    print("empieza empleado = " + str(empieza_empleado))
                if anterior_is_number and cell.value == None:
                    termina_empleado = cell.row - 1
                    termino = True
                    print("termina empleado = " + str(termina_empleado))
                    ws_empleados["C"+str(cell.row)].value = '=SUM(C'+str(empieza_empleado)+':C'+str(termina_empleado)+')'
                    ws_empleados["D"+str(cell.row)].value = '=SUM(D'+str(empieza_empleado)+':D'+str(termina_empleado)+')'
                    ws_empleados["E"+str(cell.row)].value = '=SUM(E'+str(empieza_empleado)+':E'+str(termina_empleado)+')'
                    ws_empleados["F"+str(cell.row)].value = '=E'+str(cell.row)+'-D'+str(cell.row)

                if cell.value == None:
                     anterior_is_number = False
                else:
                    anterior_is_number = True

        if not termino:
            termina_empleado = cell.row
            print("termina empleado = " + str(termina_empleado))
            ws_empleados["C"+str(cell.row+1)].value = '=SUM(C'+str(empieza_empleado)+':C'+str(termina_empleado)+')'
            ws_empleados["D"+str(cell.row+1)].value = '=SUM(D'+str(empieza_empleado)+':D'+str(termina_empleado)+')'
            ws_empleados["E"+str(cell.row+1)].value = '=SUM(E'+str(empieza_empleado)+':E'+str(termina_empleado)+')'
            ws_empleados["F"+str(cell.row+1)].value = '=E'+str(cell.row)+'-D'+str(cell.row)
        n_viaje = 1
        for viaje in workbook.get_sheet_names():
            if viaje not in ["Viajes","Empleados"]:
                n_viaje+=1
                worksheet = workbook[viaje]

                ws_viajes.cell(n_viaje,1,viaje)

                suma_row = 0
                for celda in worksheet['K']:
                    if celda.value == None:
                        break
                    suma_row = celda.row
                ws_viajes.cell(n_viaje,3,"='"+viaje+"'!K"+str(suma_row))
                ws_viajes.cell(n_viaje,2,"='"+viaje+"'!J"+str(suma_row))

                max_row = len(worksheet["A"])
                sumifs_str = "=SUMIFS('"+viaje+"'!F2:'"+viaje+"'!F"+str(max_row)+','+"'"+viaje+"'!E2:'"+viaje+"'!E"+str(max_row)+',"Transporte")'
                ws_viajes.cell(n_viaje, 4,     sumifs_str)
                sumifs_str = "=SUMIFS('"+viaje+"'!F2:'"+viaje+"'!F"+str(max_row)+','+"'"+viaje+"'!E2:'"+viaje+"'!E"+str(max_row)+',"Hospedaje")'
                ws_viajes.cell(n_viaje, 5,     sumifs_str)
                sumifs_str = "=SUMIFS('"+viaje+"'!F2:'"+viaje+"'!F"+str(max_row)+','+"'"+viaje+"'!E2:'"+viaje+"'!E"+str(max_row)+',"Alimentos")'
                ws_viajes.cell(n_viaje, 6,     sumifs_str)
                sumifs_str = "=SUMIFS('"+viaje+"'!F2:'"+viaje+"'!F"+str(max_row)+','+"'"+viaje+"'!E2:'"+viaje+"'!E"+str(max_row)+',"Otros")'
                ws_viajes.cell(n_viaje, 7,     sumifs_str)

        ws_viajes.cell(n_viaje+1,2,"=SUM(B2:B"+str(n_viaje)+")")
        ws_viajes.cell(n_viaje+1,3,"=SUM(C2:C"+str(n_viaje)+")")
        ws_viajes.cell(n_viaje+1,4,"=SUM(D2:D"+str(n_viaje)+")")
        ws_viajes.cell(n_viaje+1,5,"=SUM(E2:E"+str(n_viaje)+")")
        ws_viajes.cell(n_viaje+1,6,"=SUM(F2:F"+str(n_viaje)+")")
        ws_viajes.cell(n_viaje+1,7,"=SUM(G2:G"+str(n_viaje)+")")

        workbook.save(self.xlsx_path)
        self.ui.excel_button.setEnabled(True)
        v_row = 1
        sumaTotal = 0
        sumaTransporte = 0
        sumaHospedaje = 0
        sumaAlimentos = 0
        sumaOtros = 0
        if len(self.dicc_viajes.items()) > 4:
            self.ui.tableWidget.setRowCount(len(self.dicc_viajes.items())+2)
        for key, value in self.dicc_viajes.items():

            sumaTotal += value["Total"]
            sumaTransporte += value["Transporte"]
            sumaHospedaje += value["Hospedaje"]
            sumaAlimentos += value["Alimentos"]
            sumaOtros += value["Otros"]
            self.ui.tableWidget.setItem(v_row,0,self.esteItem(key,key))
            self.ui.tableWidget.setItem(v_row,1,self.esteItem(str(value["Total"]),""))
            self.ui.tableWidget.setItem(v_row,2,self.esteItem(str(value["Transporte"]),""))
            self.ui.tableWidget.setItem(v_row,3,self.esteItem(str(value["Hospedaje"]),""))
            self.ui.tableWidget.setItem(v_row,4,self.esteItem(str(value["Alimentos"]),""))
            self.ui.tableWidget.setItem(v_row,5,self.esteItem(str(value["Otros"]),""))
            v_row += 1
        print("total ",sumaTotal)
        print("Transporte ",sumaTransporte)
        self.ui.tableWidget.setItem(v_row,1,self.esteItem(str(sumaTotal),""))
        self.ui.tableWidget.setItem(v_row,2,self.esteItem(str(sumaTransporte),""))
        self.ui.tableWidget.setItem(v_row,3,self.esteItem(str(sumaHospedaje),""))
        self.ui.tableWidget.setItem(v_row,4,self.esteItem(str(sumaAlimentos),""))
        self.ui.tableWidget.setItem(v_row,5,self.esteItem(str(sumaOtros),""))
        self.ui.textBrowser.append("Resumen: "+self.xlsx_path)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    app.setWindowIcon(QIcon(join(window.dirPath,'logo.ico')))
    window.setWindowIcon(QIcon(join(window.dirPath,'logo.ico')))
    sys.exit(app.exec_())
