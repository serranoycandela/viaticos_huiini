#-*- encoding: utf-8 -*-
#import xml.etree.ElementTree as etree
import os
#import xlsxwriter
import sys
from PySide2.QtCore import Qt
#from PySide2 import QtGui, QtCore, QtWidgets
from PySide2.QtWidgets import QApplication, QMainWindow, QInputDialog, QFileDialog, QTableWidgetItem, QListView, QMessageBox, QLineEdit
#from PySide2.QtCore import QFile, QRect
from PySide2.QtGui import QIcon
from gui import Ui_MainWindow
import json
from os.path import dirname, realpath, join, abspath
#from openpyxl.styles import Alignment

from openpyxl import load_workbook#,  Workbook
#from openpyxl.utils import get_column_letter
#from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.datavalidation import DataValidation
from FacturasLocal import FacturaLocal as Factura
import pandas as pd
from datetime import datetime

# pyinstaller.exe --noconsole .\viajero_helper.py
# pyside2-uic gui.ui -o gui.py

if getattr(sys, 'frozen', False):
    # we are running in a bundle
    scriptDirectory = os.path.dirname(sys.executable)
    appDataDirectory = os.path.expandvars('%APPDATA%\huiini')
else:
    # we are running in a normal Python environment
    scriptDirectory = os.path.dirname(os.path.abspath(__file__))
    appDataDirectory = join(scriptDirectory,"huiini_aux_files")

class MainWindow(QMainWindow):
    def __init__(self):
        super(MainWindow, self).__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)

        
        self.ui.carpeta_personal.clicked.connect(self.escoje_carpeta_de_viaje)
        self.ui.excel_button.clicked.connect(self.abre_excel)

        self.ui.tableWidget.setColumnCount(6)
        self.ui.tableWidget.setColumnWidth(0,120)#viaje
        self.ui.tableWidget.setColumnWidth(1,70)#total
        self.ui.tableWidget.setColumnWidth(2,70)#transporte
        self.ui.tableWidget.setColumnWidth(3,70)#hospedaje
        self.ui.tableWidget.setColumnWidth(4,70)#alimentos
        self.ui.tableWidget.setColumnWidth(5,70)#otros
        #header = self.tableWidget.verticalHeader()
        self.ponEncabezado()
        self.dicc_users = {}
        self.dicc_viajes = {}
        if len(sys.argv) > 1:
            self.ui.carpeta_personal.setEnabled(False)


    def esteItem(self, text, tooltip):
        item = QTableWidgetItem(text)
        item.setToolTip(tooltip)
        item.setFlags(item.flags() ^ Qt.ItemIsEditable)
        return item
    def tipo_de_gasto(self, clave_ps):
        tipo = "Otros"
        if clave_ps.startswith('7811') or clave_ps.startswith('9511') or clave_ps.startswith('1510') or clave_ps.startswith('1511') or clave_ps.startswith('1512'):
            tipo = "Transporte"
        if clave_ps.startswith('9010'):
            tipo = "Alimentos"
        if clave_ps.startswith('9011'):
            tipo = "Hospedaje"

        return(tipo)


    def abre_excel(self):
        try:
            os.startfile(self.xlsx_path)
        except:
            QMessageBox.information(self, "Information", "El sistema no tiene una aplicación por default para abrir exceles" )

        
    def escoje_carpeta_de_viaje(self):    
        file_dialog = QFileDialog()
        file_dialog.setFileMode(QFileDialog.DirectoryOnly)
        file_dialog.setOption(QFileDialog.DontUseNativeDialog, True)
        file_view = file_dialog.findChild(QListView, 'listView')
        if file_dialog.exec():
            trip_folder = file_dialog.selectedFiles()[0]
            self.prellenar_formato(trip_folder)
        
    def get_metadata(self):
        
        if os.path.isfile(os.path.join(appDataDirectory,"proyecto.txt")):
            with open(os.path.join(appDataDirectory,"proyecto.txt"), 'r') as f:
                proyecto_guardado = f.readline()
        else:
            proyecto_guardado = False

        if os.path.isfile(os.path.join(appDataDirectory,"viajero.txt")):
            with open(os.path.join(appDataDirectory,"viajero.txt"), 'r') as f:
                viajero_guardado = f.readline()
        else:
            viajero_guardado = "Ingresa tu RFC"
        
        proyectos_sheet_id = "1a3-O0OvzDQect8EszxeeDrKbrLxNaw2Lti7t37Usfq8"
        r = "https://docs.google.com/spreadsheets/export?id={}&exportFormat=csv".format(proyectos_sheet_id)
        df_proyectos = pd.read_csv(r)
        lista_proyectos = df_proyectos['NOMBRE'].tolist()
        if proyecto_guardado:
            id_proyecto = lista_proyectos.index(proyecto_guardado)
        else:
            id_proyecto = 0
        
        proyecto, ok = QInputDialog().getItem(self, "Proyecto",
                                            "Proyecto:", lista_proyectos, id_proyecto, False)
        if ok and proyecto:
            with open(os.path.join(appDataDirectory,"proyecto.txt"), "w") as f:
                f.write(proyecto)

        viajeros_sheet_id = "14_DmyXmBJx7eof65VelFrkrGKK4Rw6ZSR0nRSFFiio4"
        r = "https://docs.google.com/spreadsheets/export?id={}&exportFormat=csv".format(viajeros_sheet_id)
        df_viajeros = pd.read_csv(r)
        
        viajero_rfc, ok = QInputDialog().getText(self, "RFC",
                                     "RFC:", QLineEdit.Normal,
                                     viajero_guardado)
        if ok and viajero_rfc:
            viejero_row = df_viajeros[df_viajeros['RFC'] == viajero_rfc] 
            viajero = viejero_row['NOMBRE'].loc[viejero_row.index[0]]
            coordinador = viejero_row['COORDINADOR'].loc[viejero_row.index[0]]
            empresa = viejero_row['EMPRESA'].loc[viejero_row.index[0]]
            with open(os.path.join(appDataDirectory,"viajero.txt"), "w") as f:
                f.write(viajero_rfc)

        # viajero, ok = QInputDialog().getItem(self, "Nombre",
        #                                     "Nombre:", lista_viajeros, id_viajero, False)
        # if ok and viajero:
            # with open(os.path.join(appDataDirectory,"viajero.txt"), "w") as f:
            #     f.write(viajero)


        return [viajero, coordinador, proyecto, empresa]
        


    def prellenar_formato(self, trip_folder):

        try:
            
            metadatos = self.get_metadata()
            viajero = metadatos[0]
            coordinador = metadatos[1]
            proyecto = metadatos[2]
            empresa = metadatos[3]
        except:
            QMessageBox.information(self, "Information", "Problema de conexión" )
            return 0

        self.ui.textBrowser.clear()
        self.ui.tableWidget.clear()
        self.ui.tableWidget.repaint()
        self.ponEncabezado()
        self.ui.tableWidget.setRowCount(6)
        self.ui.tableWidget.repaint()

        wb_template = load_workbook(os.path.join(appDataDirectory,"FORMATO_template_"+empresa+".xlsx"))
        ws_template = wb_template[wb_template.get_sheet_names()[0]]

        row_f = 11
        #folder = os.path.split(path)[1]
        por_tipo={"Total":0,"Transporte":0,"Alimentos":0,"Hospedaje":0,"Otros":0}
        hay_facturas = False
        date_format = '%Y-%m-%d'
        fechas = []
        for archivo in os.listdir(trip_folder):
            if archivo.endswith(".xml"):
                row_f += 1
                xml_path = os.path.join(trip_folder,archivo)
                factura = Factura(xml_path)
                total = factura.total
                iva = factura.traslados["IVA"]["importe"]
                uuid = factura.UUID
                fecha = factura.fechaTimbrado.split("T")[0]
                fechas.append(datetime.strptime(fecha, date_format))
                provedor = factura.EmisorNombre
                clave_ps = factura.conceptos[0]['clave_concepto']
                tipo = self.tipo_de_gasto(clave_ps)
                ws_template.cell(row_f, 1,    fecha)#fecha
                ws_template.cell(row_f, 2,    uuid)#fecha
                ws_template.cell(row_f, 3,    tipo)#tipo
                ws_template.cell(row_f, 4,    provedor)#provedor
                ws_template.cell(row_f, 5,    factura.subTotal )#importe
                ws_template.cell(row_f, 6,    iva)#traslado_iva
                ws_template.cell(row_f, 7,    total)
                por_tipo[tipo]+=total
                por_tipo["Total"]+=total
                hay_facturas = True

        ws_template.cell(8, 1, viajero)#nombre
        ws_template.cell(8, 3, proyecto)#proyecto
        ws_template.cell(57, 4, viajero)#firma
        ws_template.cell(57, 1, coordinador)#coordinador
        
        if hay_facturas:
            fechas.sort(reverse=False)
            fecha_inicio = fechas[0]
            fecha_fin = fechas[len(fechas)-1]
            if fecha_inicio == fecha_fin:
                fechas_str = fecha_fin.strftime('%d-%m-%Y')
            else:
                fechas_str = fecha_inicio.strftime('%d-%m-%Y') + "_" + fecha_fin.strftime('%d-%m-%Y')
            
            self.xlsx_path = os.path.join(trip_folder,"FORMATO_"+viajero.replace(" ","-")+"_"+fechas_str+".xlsx")
            self.ui.textBrowser.append("Procesando: "+trip_folder)
            self.ui.tableWidget.setItem(1,0,self.esteItem(fechas_str,""))
            self.ui.tableWidget.setItem(1,1,self.esteItem(str(por_tipo["Total"]),""))
            self.ui.tableWidget.setItem(1,2,self.esteItem(str(por_tipo["Transporte"]),""))
            self.ui.tableWidget.setItem(1,3,self.esteItem(str(por_tipo["Hospedaje"]),""))
            self.ui.tableWidget.setItem(1,4,self.esteItem(str(por_tipo["Alimentos"]),""))
            self.ui.tableWidget.setItem(1,5,self.esteItem(str(por_tipo["Otros"]),""))
            
            ws_template.cell(4, 6, fecha_inicio.strftime('%d-%m-%Y'))#fecha_inicio
            ws_template.cell(5, 6, fecha_fin.strftime('%d-%m-%Y'))#fecha_fin
            self.ui.textBrowser.append("Formato de gastos de viaje creado en: "+self.xlsx_path)
            wb_template.save(self.xlsx_path)
            self.ui.excel_button.setEnabled(True)
        else:
            QMessageBox.information(self, "Information", "No hay facturas en este folder" )

        if len(sys.argv) > 1:
            self.abre_excel()
            sys.exit()

    def ponEncabezado(self):
        self.ui.tableWidget.setHorizontalHeaderItem (0, QTableWidgetItem("Viaje"))
        self.ui.tableWidget.setHorizontalHeaderItem (1, QTableWidgetItem("Total"))
        self.ui.tableWidget.setHorizontalHeaderItem (2, QTableWidgetItem("Transporte"))
        self.ui.tableWidget.setHorizontalHeaderItem (3, QTableWidgetItem("hospedaje"))
        self.ui.tableWidget.setHorizontalHeaderItem (4, QTableWidgetItem("Alimentos"))
        self.ui.tableWidget.setHorizontalHeaderItem (5, QTableWidgetItem("Otros"))

  

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    if len(sys.argv) > 1:
        window.prellenar_formato(sys.argv[1])
        
    else:
        window.show()
        app.setWindowIcon(QIcon(join(scriptDirectory,'logo.ico')))
        window.setWindowIcon(QIcon(join(scriptDirectory,'logo.ico')))
    sys.exit(app.exec_())