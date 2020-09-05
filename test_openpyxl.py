from openpyxl import load_workbook,  Workbook

wb = load_workbook('C:\\Users\\arabela\\Documents\\GitHub\\viaticos_huiini\\resumen.xlsx')
ws1 = wb.create_sheet("Mysheet7")
ws1.cell(1,1,"huevos")
print(wb.sheetnames)
wb.save('C:\\Users\\arabela\\Documents\\GitHub\\viaticos_huiini\\resumen.xlsx')
wb = Workbook()
ws1 = wb.create_sheet("Mysheet")
ws1.cell(row=1,column=1,value="huevos")
wb.save('C:\\Users\\arabela\\Documents\\GitHub\\viaticos_huiini\\resumen2.xlsx')
xlsx_path = 'C:\\Users\\arabela\\Documents\\GitHub\\viaticos_huiini\\datos_prueba\\AGOSTO\\resumen_AGOSTO.xlsx'
data = load_workbook(xlsx_path, data_only=True)
print(data.get_sheet_names())
viajes = data["Viajes"]
celda = viajes.cell(2,4)
valor = celda.value
print("------------------------"+str(valor))
import os
path = os.getenv('LOCALAPPDATA')
array = os.listdir(path)
print(path)

pdflatex_path = os.path.join(path,"Programs","MiKTeX","miktex","bin","x64","pdflatex.exe")
pdflatex_path = pdflatex_path.replace("\\","\\\\")
print(pdflatex_path)
print("AAAAAA".title())
